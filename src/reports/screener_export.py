"""screener_output.xlsx generator — Sprint 3, Day 17.

Spec Module 3 (3.5) / Deliverables D-07/D-08: one sheet per preset (6
total), 20+ KPIs, colour-coded by threshold, sorted by each preset's
rank_by column.

Design decision (documented, not silently assumed): each sheet contains
only QUALIFYING companies (matches D-07's "screener results" language,
not the full 92-company universe). Colour-coding is scoped to each
preset's own filter-criteria columns (green, since qualifying rows
inherently pass — logic is written generically off config so it would
correctly render red too, if non-qualifying rows were ever included in
future), plus an independent green/yellow/red tier on the composite
score column (>=70 / 40-70 / <40) for additional visual signal beyond
pass/fail. Flagged for team lead review as an interpretation of the
spec's cell-colouring intent.
"""

from __future__ import annotations

import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.screener.composite_score import compute_sector_relative_composite_score
from src.screener.engine import load_config, load_screener_universe, run_custom_screen
from src.screener.turnaround import run_turnaround_watch

DB_PATH = "data/nifty100.db"
OUTPUT_PATH = "output/screener_output.xlsx"

THRESHOLD_PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

DISPLAY_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "composite_score_sector_relative",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "asset_turnover",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "market_cap_crore",
    "sales",
    "net_profit",
]

_OPS_LABEL = {">": "min", ">=": "min", "<": "max", "<=": "max", "==": "exact"}


def _meets_filter(value: float | None, op: str, threshold: float) -> bool | None:
    """Re-evaluate a single filter condition for colour-coding purposes.
    Returns None if value is missing (no colour applied)."""
    if value is None or pd.isna(value):
        return None
    if op == ">":
        return value > threshold
    if op == ">=":
        return value >= threshold
    if op == "<":
        return value < threshold
    if op == "<=":
        return value <= threshold
    if op == "==":
        return value == threshold
    return None


def _composite_score_fill(score: float | None) -> PatternFill | None:
    if score is None or pd.isna(score):
        return None
    if score >= 70:
        return GREEN_FILL
    if score >= 40:
        return YELLOW_FILL
    return RED_FILL


def _write_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    preset_config: dict,
    metrics_map: dict,
) -> None:
    """Write one preset's results to a worksheet with headers, data, and
    colour-coding."""
    columns = [c for c in DISPLAY_COLUMNS if c in df.columns]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Map: real dataframe column -> (op, threshold) from this preset's
    # own filters, so we can colour-code the exact columns this preset
    # actually screens on.
    filter_col_conditions: dict[str, tuple[str, float]] = {}
    for f in preset_config.get("filters", []):
        real_col = metrics_map.get(f["metric"], f["metric"])
        filter_col_conditions[real_col] = (f["op"], f["value"])

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name == "composite_score_sector_relative":
                fill = _composite_score_fill(value)
                if fill:
                    cell.fill = fill
            elif col_name in filter_col_conditions:
                op, threshold = filter_col_conditions[col_name]
                meets = _meets_filter(value, op, threshold)
                if meets is True:
                    cell.fill = GREEN_FILL
                elif meets is False:
                    cell.fill = RED_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            14, len(col_name) + 2
        )

    ws.freeze_panes = "A2"


def generate_screener_output() -> None:
    """Generate screener output."""
    conn = sqlite3.connect(DB_PATH)
    config = load_config()

    universe = load_screener_universe(conn)

    # Day 17 design decision: compute_sector_relative_composite_score()
    # masks return_on_equity_pct/return_on_capital_employed_pct to NaN for
    # companies tripping Day 13's sanity bound (HAL, BEL, INDIGO,
    # ICICIPRULI, HDFCLIFE -- confirmed via day17_diagnose_sanity_bound.py).
    # Reassigning `universe` to that function's return value (rather than
    # a separate variable) is DELIBERATE here: the same masked ROE/ROCE
    # then flows into every preset's filters below, not just the
    # composite score. An implausible ROE (e.g. INDIGO's 892.6%) shouldn't
    # pass a "ROE > 12%" filter any more than it should be averaged into a
    # score -- both consume the same untrustworthy balance-sheet data.
    # Confirmed via day17_diagnose_preset_shift.py: this drops INDIGO from
    # quality_compounder (22->21) and debt_free_blue_chip (20->19),
    # resolving the exact issue flagged as a "KNOWN ISSUE carried to
    # Day 17" in config/screener_config.yaml's debt_free_blue_chip preset.
    universe = compute_sector_relative_composite_score(universe, conn)

    companies = pd.read_sql_query(
        "SELECT id AS company_id, company_name FROM companies", conn
    )
    universe = universe.merge(companies, on="company_id", how="left")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    for preset_name in THRESHOLD_PRESETS:
        preset = config["presets"][preset_name]
        result = run_custom_screen(
            universe,
            preset["filters"],
            preset["rank_by"],
            config,
            exclude_sectors=preset.get("exclude_sectors"),
        )
        ws = wb.create_sheet(title=preset_name[:31])  # Excel sheet name limit
        _write_sheet(ws, result, preset, config["metrics"])
        print(f"{preset_name}: {len(result)} companies written")

    turnaround_result = run_turnaround_watch(universe, conn)
    ws = wb.create_sheet(title="turnaround_watch")
    turnaround_preset_config = {
        "filters": []
    }  # no simple threshold filters to colour-code
    _write_sheet(ws, turnaround_result, turnaround_preset_config, config["metrics"])
    print(f"turnaround_watch: {len(turnaround_result)} companies written")

    import os

    os.makedirs("output", exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")

    conn.close()


if __name__ == "__main__":
    generate_screener_output()
