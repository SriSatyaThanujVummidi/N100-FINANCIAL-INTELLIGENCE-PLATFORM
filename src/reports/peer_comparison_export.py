"""peer_comparison.xlsx generator — Sprint 3, Day 20.

Spec Module 9 / Deliverable D-09: one sheet per peer group (11 total),
company_id + company_name + 20 metric columns + percentile rank per
metric, colour-coded (green >=75th pctile, yellow 25-75th, red <=25th),
benchmark row highlighted gold, summary row with group medians.

Reuses Day 17's compute_sector_relative_composite_score() (masked
ROE/ROCE/composite for HAL/BEL/INDIGO/ICICIPRULI/HDFCLIFE) and Day 18's
peer_percentiles table, so raw values and percentiles stay consistent —
never showing a real implausible ratio next to a blank percentile.

Design decision: summary row median is computed only for the 20 raw
metric columns, not the percentile columns (a "median percentile" isn't
a meaningful concept the way a median ROE is) — left blank there.
"""

from __future__ import annotations

import os
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.composite_score import compute_sector_relative_composite_score
from src.screener.engine import load_screener_universe

DB_PATH = "data/nifty100.db"
OUTPUT_PATH = "output/peer_comparison.xlsx"

METRIC_COLUMNS = [
    "broad_sector",
    "composite_score_sector_relative",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "capex_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "asset_turnover",
    "return_on_assets_pct",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "market_cap_crore",
    "sales",
]

PERCENTILE_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_FONT = Font(italic=True, bold=True)


def build_peer_comparison_universe(conn: sqlite3.Connection) -> pd.DataFrame:
    """Universe restricted to the 56 companies in a real peer group,
    with masked ROE/ROCE/composite (Day 17) and pivoted percentiles
    (Day 18) merged in wide format."""
    universe = load_screener_universe(conn)
    universe = compute_sector_relative_composite_score(universe, conn)

    peer_groups = pd.read_sql_query(
        "SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn
    )
    universe = universe.merge(peer_groups, on="company_id", how="inner")  # only the 56

    companies = pd.read_sql_query(
        "SELECT id AS company_id, company_name FROM companies", conn
    )
    universe = universe.merge(companies, on="company_id", how="left")

    percentiles = pd.read_sql_query(
        "SELECT company_id, metric, percentile_rank FROM peer_percentiles", conn
    )
    wide = percentiles.pivot(
        index="company_id", columns="metric", values="percentile_rank"
    )
    wide.columns = [f"{c}_percentile" for c in wide.columns]
    universe = universe.merge(wide, on="company_id", how="left")

    return universe


def _percentile_fill(pctile: float | None) -> PatternFill | None:
    if pctile is None or pd.isna(pctile):
        return None
    if pctile >= 0.75:
        return GREEN_FILL
    if pctile <= 0.25:
        return RED_FILL
    return YELLOW_FILL


def _write_group_sheet(wb: Workbook, group_name: str, group_df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=group_name[:31])

    columns = (
        ["company_id", "company_name"]
        + METRIC_COLUMNS
        + [f"{m}_percentile" for m in PERCENTILE_METRICS]
    )
    columns = [
        c
        for c in columns
        if c in group_df.columns or c in ("company_id", "company_name")
    ]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    group_df = group_df.sort_values("is_benchmark", ascending=False)

    row_idx = 2
    for _, row in group_df.iterrows():
        is_benchmark = bool(row.get("is_benchmark"))
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if is_benchmark:
                cell.fill = GOLD_FILL
            elif col_name.endswith("_percentile"):
                fill = _percentile_fill(value)
                if fill:
                    cell.fill = fill
        row_idx += 1

    # --- summary row: median for the 20 raw metric columns only ---
    summary_row = row_idx
    ws.cell(row=summary_row, column=1, value="MEDIAN").font = SUMMARY_FONT
    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in ("company_id", "company_name"):
            continue
        if col_name.endswith("_percentile"):
            continue  # documented: median percentile not meaningful, left blank
        if col_name == "broad_sector":
            continue  # categorical, no median
        if col_name in group_df.columns:
            median_val = group_df[col_name].median()
            cell = ws.cell(row=summary_row, column=col_idx, value=median_val)
            cell.font = SUMMARY_FONT

    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            14, len(col_name) + 2
        )
    ws.freeze_panes = "A2"


def generate_peer_comparison_output() -> None:
    """Generate peer comparison output."""
    conn = sqlite3.connect(DB_PATH)
    universe = build_peer_comparison_universe(conn)

    wb = Workbook()
    wb.remove(wb.active)

    for group_name, group_df in universe.groupby("peer_group_name"):
        _write_group_sheet(wb, group_name, group_df)
        print(f"{group_name}: {len(group_df)} companies written")

    os.makedirs("output", exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(
        f"\nSaved: {OUTPUT_PATH} ({len(universe['peer_group_name'].unique())} sheets)"
    )

    conn.close()


if __name__ == "__main__":
    generate_peer_comparison_output()
