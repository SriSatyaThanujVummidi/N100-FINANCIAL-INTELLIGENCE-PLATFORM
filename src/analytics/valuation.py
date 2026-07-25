"""Module 6: Valuation & Market Data Module — Day 26.

Computes FCF yield, sector-relative P/E positioning, and overvaluation/
discount flags for all 92 companies using market_cap.xlsx (P/E, P/B,
EV/EBITDA, simulated) joined with financial_ratios (FCF, real/computed).

JUDGMENT CALL (documented, not silently applied): financial_ratios is
indexed by fiscal year (e.g. '2024-03'), market_cap by calendar year
(e.g. 2024) — these aren't the same axis, and an exact-year join would
silently drop non-March-FYE companies (SIEMENS=Sep, NESTLEIND/AMBUJACEM/
EICHERMOT/BOSCHLTD/ABB=Dec, per Day 15's finding), repeating that bug.
Instead, FCF and market multiples are each taken from their OWN latest
available row per company, independently. Flag for team lead review if
exact fiscal-to-calendar alignment is required.
"""

from __future__ import annotations

import logging
import os
import sqlite3
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(_PROJECT_ROOT / ".env")

_DB_PATH_RAW = os.getenv("DB_PATH", "data/nifty100.db")
_DB_PATH = Path(_DB_PATH_RAW)
if not _DB_PATH.is_absolute():
    _DB_PATH = _PROJECT_ROOT / _DB_PATH

_OUTPUT_DIR = _PROJECT_ROOT / "output"
_OUTPUT_DIR.mkdir(exist_ok=True)

CAUTION_MULTIPLIER = 1.5
DISCOUNT_MULTIPLIER = 0.7

# Mirrors Day 13's Option B sanity bound — a P/E computed off a near-zero
# or negative earnings base can produce nonsense multiples the same way
# HAL/BEL's near-zero equity produced nonsense ROE. Bound chosen generously
# (spec Section 28's widest sector P/E band tops out at 70x) so only clear
# data artifacts are excluded, not legitimately expensive growth stocks.
PE_SANITY_BOUND = 200.0


def _get_connection() -> sqlite3.Connection:
    if not _DB_PATH.exists():
        raise FileNotFoundError(f"Database not found at {_DB_PATH}")
    conn = sqlite3.connect(str(_DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def load_latest_market_cap() -> pd.DataFrame:
    """One market_cap row per company at their latest available calendar year."""
    query = """
        SELECT m.*
        FROM market_cap m
        INNER JOIN (
            SELECT company_id, MAX(year) AS max_year
            FROM market_cap
            GROUP BY company_id
        ) latest
        ON m.company_id = latest.company_id AND m.year = latest.max_year
    """
    with _get_connection() as conn:
        return pd.read_sql_query(query, conn)


def load_all_market_cap() -> pd.DataFrame:
    """Full market_cap history — needed for the 5yr median P/E column."""
    with _get_connection() as conn:
        return pd.read_sql_query("SELECT * FROM market_cap", conn)


def load_latest_fcf() -> pd.DataFrame:
    """One financial_ratios row (FCF only) per company at their own latest
    reported fiscal year."""
    query = """
        SELECT r.company_id, r.year AS fiscal_year, r.free_cash_flow_cr
        FROM financial_ratios r
        INNER JOIN (
            SELECT company_id, MAX(year) AS max_year
            FROM financial_ratios
            GROUP BY company_id
        ) latest
        ON r.company_id = latest.company_id AND r.year = latest.max_year
    """
    with _get_connection() as conn:
        return pd.read_sql_query(query, conn)


def load_companies_with_sector() -> pd.DataFrame:
    """Load companies with sector."""
    query = """
        SELECT c.id AS company_id, c.company_name, s.broad_sector
        FROM companies c
        LEFT JOIN sectors s ON c.id = s.company_id
    """
    with _get_connection() as conn:
        return pd.read_sql_query(query, conn)


def compute_5yr_median_pe(all_mcap: pd.DataFrame) -> pd.DataFrame:
    """Median P/E per company over their most recent up-to-5 reported years.
    Uses fewer years if a company's market_cap history is shorter than 5yr
    (e.g. any company added to the universe partway through 2019-2024)."""
    results = []
    for company_id, group in all_mcap.groupby("company_id"):
        recent = group.sort_values("year").tail(5)
        pe_clean = recent["pe_ratio"].where(recent["pe_ratio"].abs() <= PE_SANITY_BOUND)
        median_pe = pe_clean.dropna().median()
        results.append({"company_id": company_id, "median_pe_5yr": median_pe})
    return pd.DataFrame(results)


def compute_sector_median_pe(
    latest_mcap: pd.DataFrame, companies_sector: pd.DataFrame
) -> pd.DataFrame:
    """Sector median P/E in the latest year, using the sanity-bounded P/E
    only — one implausible P/E in a sector shouldn't skew the whole
    sector's benchmark, same principle as Day 17's ROE/ROCE masking
    protecting composite score percentile windows."""
    merged = latest_mcap.merge(companies_sector, on="company_id", how="left")
    merged["pe_clean"] = merged["pe_ratio"].where(
        merged["pe_ratio"].abs() <= PE_SANITY_BOUND
    )
    sector_medians = (
        merged.dropna(subset=["pe_clean", "broad_sector"])
        .groupby("broad_sector")["pe_clean"]
        .median()
        .reset_index()
        .rename(columns={"pe_clean": "sector_median_pe"})
    )
    return sector_medians


def classify_valuation(pe: Optional[float], sector_median_pe: Optional[float]) -> str:
    """Caution if P/E > sector_median x 1.5. Discount if P/E < sector_median x 0.7.
    Otherwise Fair. N/A if either input is missing or P/E fails the sanity bound."""
    if (
        pe is None
        or pd.isna(pe)
        or sector_median_pe is None
        or pd.isna(sector_median_pe)
    ):
        return "N/A"
    if abs(pe) > PE_SANITY_BOUND:
        return "N/A"
    if sector_median_pe == 0:
        return "N/A"
    if pe > sector_median_pe * CAUTION_MULTIPLIER:
        return "Caution"
    if pe < sector_median_pe * DISCOUNT_MULTIPLIER:
        return "Discount"
    return "Fair"


def build_valuation_summary() -> pd.DataFrame:
    """Build valuation summary."""
    logger.info("Loading data...")
    latest_mcap = load_latest_market_cap()
    all_mcap = load_all_market_cap()
    latest_fcf = load_latest_fcf()
    companies_sector = load_companies_with_sector()

    logger.info("Computing 5yr median P/E per company...")
    median_pe_5yr = compute_5yr_median_pe(all_mcap)

    logger.info("Computing sector median P/E (latest year, sanity-bounded)...")
    sector_medians = compute_sector_median_pe(latest_mcap, companies_sector)

    df = companies_sector.merge(latest_mcap, on="company_id", how="left")
    df = df.merge(median_pe_5yr, on="company_id", how="left")
    df = df.merge(latest_fcf, on="company_id", how="left")
    df = df.merge(sector_medians, on="broad_sector", how="left")

    # FCF Yield = FCF / market_cap_crore x 100
    df["fcf_yield_pct"] = df.apply(
        lambda r: (
            (r["free_cash_flow_cr"] / r["market_cap_crore"] * 100)
            if pd.notna(r.get("free_cash_flow_cr"))
            and pd.notna(r.get("market_cap_crore"))
            and r.get("market_cap_crore") not in (0, None)
            else None
        ),
        axis=1,
    )

    # PE vs sector median, as a percentage difference
    df["pe_vs_sector_median_pct"] = df.apply(
        lambda r: (
            ((r["pe_ratio"] / r["sector_median_pe"]) - 1) * 100
            if pd.notna(r.get("pe_ratio"))
            and pd.notna(r.get("sector_median_pe"))
            and r.get("sector_median_pe") not in (0, None)
            else None
        ),
        axis=1,
    )

    df["flag"] = df.apply(
        lambda r: classify_valuation(r.get("pe_ratio"), r.get("sector_median_pe")),
        axis=1,
    )

    out = df.rename(
        columns={
            "pe_ratio": "pe",
            "pb_ratio": "pb",
            "ev_ebitda": "ev_ebitda",
            "median_pe_5yr": "median_pe_5yr",
        }
    )[
        [
            "company_id",
            "company_name",
            "broad_sector",
            "pe",
            "pb",
            "ev_ebitda",
            "fcf_yield_pct",
            "median_pe_5yr",
            "pe_vs_sector_median_pct",
            "flag",
        ]
    ].rename(
        columns={
            "broad_sector": "sector",
            "pe": "P/E",
            "pb": "P/B",
            "ev_ebitda": "EV/EBITDA",
            "fcf_yield_pct": "FCF_yield_pct",
            "median_pe_5yr": "5yr_median_PE",
            "pe_vs_sector_median_pct": "PE_vs_sector_median_pct",
        }
    )

    out = out.sort_values("company_id").reset_index(drop=True)
    return out


def _apply_flag_colour(ws, flag_col_idx: int, n_rows: int) -> None:
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row_idx in range(2, n_rows + 2):
        cell = ws.cell(row=row_idx, column=flag_col_idx)
        if cell.value == "Fair":
            cell.fill = green
        elif cell.value == "Discount":
            cell.fill = yellow
        elif cell.value == "Caution":
            cell.fill = red


def export_valuation_summary(df: pd.DataFrame, path: Path) -> None:
    """Export valuation summary."""
    df.to_excel(path, index=False, sheet_name="Valuation Summary")

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["Valuation Summary"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    flag_col_idx = df.columns.get_loc("flag") + 1 if "flag" in df.columns else None
    if flag_col_idx:
        _apply_flag_colour(ws, flag_col_idx, len(df))

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).map(len).max() if len(df) else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    wb.save(path)
    logger.info(f"Wrote {path} ({len(df)} rows)")


def export_valuation_flags(df: pd.DataFrame, path: Path) -> None:
    """Export valuation flags."""
    flagged = df[df["flag"].isin(["Caution", "Discount"])].copy()
    flagged.to_csv(path, index=False)
    logger.info(f"Wrote {path} ({len(flagged)} flagged rows)")


def main() -> None:
    """Main."""
    df = build_valuation_summary()

    n_total = len(df)
    n_na_flag = int((df["flag"] == "N/A").sum())
    n_fair = int((df["flag"] == "Fair").sum())
    n_caution = int((df["flag"] == "Caution").sum())
    n_discount = int((df["flag"] == "Discount").sum())

    logger.info(
        f"Valuation summary: {n_total} companies total | "
        f"Fair={n_fair} Caution={n_caution} Discount={n_discount} N/A={n_na_flag}"
    )
    if n_total != 92:
        logger.warning(f"Expected 92 companies (AC-01), got {n_total}.")

    export_valuation_summary(df, _OUTPUT_DIR / "valuation_summary.xlsx")
    export_valuation_flags(df, _OUTPUT_DIR / "valuation_flags.csv")


if __name__ == "__main__":
    main()
