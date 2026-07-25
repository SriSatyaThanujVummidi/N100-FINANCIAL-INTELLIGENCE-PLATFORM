"""
src/analytics/day32_capital_allocation_report.py

Day 32 — Sprint 5, Module 7 (Capital Allocation Report)

Tasks per spec:
  1. Verify output/capital_allocation.csv (Sprint 2 Day 11) is complete
     for all 92 companies x all years.
  2. Distribution summary: count of companies per pattern, latest year.
  3. capital_allocation_label column in cashflow_intelligence.xlsx —
     ALREADY PRESENT as of Day 31 (reused Day 11's CSV directly), so
     nothing new to build here; this script just re-confirms it.
  4. output/pattern_changes.csv — companies whose pattern changed
     year-over-year (e.g. Reinvestor -> Distress Signal).

Output:
    output/capital_allocation_distribution.csv  - pattern_label, company_count (latest year)
    output/pattern_changes.csv                  - company_id, from_year, to_year, from_pattern, to_pattern

=== Real-data finding under investigation ===
First run: 532 pattern changes across 92 companies (971 possible
year-to-year transitions) -> ~55% average change rate. Added
diagnose_change_frequency() to check whether this is spread evenly
across companies (possible classifier sensitivity near sign boundaries,
tying back to Day 11's documented "zero-value cash flows treated as +"
convention) or concentrated in a genuinely volatile subset (a real
finding, not a bug) before accepting the number.
"""

import csv
import logging
import sqlite3
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"
CAPITAL_ALLOC_CSV = OUTPUT_DIR / "capital_allocation.csv"
CASHFLOW_XLSX = OUTPUT_DIR / "cashflow_intelligence.xlsx"

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


def load_capital_allocation() -> list[dict]:
    """Load capital allocation."""
    if not CAPITAL_ALLOC_CSV.exists():
        raise FileNotFoundError(
            f"{CAPITAL_ALLOC_CSV} not found — expected from Sprint 2 Day 11's "
            "generate_capital_allocation.py"
        )
    with open(CAPITAL_ALLOC_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def verify_coverage(conn: sqlite3.Connection, rows: list[dict]) -> dict:
    """Coverage check: every (company_id, year) that has a cashflow row
    should have a matching row in capital_allocation.csv, since the
    pattern classifier needs CFO/CFI/CFF signs to run at all."""
    company_ids = sorted(r["id"] for r in conn.execute("SELECT id FROM companies"))
    cf_year_pairs = {
        (r["company_id"], r["year"])
        for r in conn.execute("SELECT company_id, year FROM cashflow")
    }
    alloc_pairs = {(r["company_id"], r["year"]) for r in rows}
    alloc_companies = {r["company_id"] for r in rows}

    missing_companies = [c for c in company_ids if c not in alloc_companies]
    missing_pairs = sorted(cf_year_pairs - alloc_pairs)

    return {
        "total_companies_in_db": len(company_ids),
        "companies_in_capital_allocation": len(alloc_companies),
        "missing_companies": missing_companies,
        "cashflow_rows": len(cf_year_pairs),
        "capital_allocation_rows": len(alloc_pairs),
        "missing_company_year_pairs": missing_pairs,
    }


def latest_year_distribution(rows: list[dict]) -> list[dict]:
    """For each company, take its latest reported year's pattern_label,
    then count companies per label — mirrors the spec's 'latest year'
    framing without assuming every company shares the same fiscal-year
    string (Day 15's SIEMENS/NESTLEIND lesson)."""
    latest_by_company: dict[str, tuple[str, str]] = {}
    for r in rows:
        cid, year, label = r["company_id"], r["year"], r["pattern_label"]
        if cid not in latest_by_company or year > latest_by_company[cid][0]:
            latest_by_company[cid] = (year, label)

    counts: dict[str, int] = {}
    for _, label in latest_by_company.values():
        counts[label] = counts.get(label, 0) + 1

    return [
        {"pattern_label": label, "company_count": count}
        for label, count in sorted(counts.items(), key=lambda x: -x[1])
    ]


def detect_pattern_changes(rows: list[dict]) -> list[dict]:
    """For each company, walk its years in order and record every
    consecutive-year transition where pattern_label changed."""
    by_company: dict[str, list[tuple[str, str]]] = {}
    for r in rows:
        by_company.setdefault(r["company_id"], []).append(
            (r["year"], r["pattern_label"])
        )

    changes = []
    for cid, year_labels in by_company.items():
        year_labels.sort(key=lambda x: x[0])
        for (y1, l1), (y2, l2) in zip(year_labels, year_labels[1:]):
            if l1 != l2:
                changes.append(
                    {
                        "company_id": cid,
                        "from_year": y1,
                        "to_year": y2,
                        "from_pattern": l1,
                        "to_pattern": l2,
                    }
                )
    return changes


def diagnose_change_frequency(rows: list[dict]) -> None:
    """Checks whether the observed pattern changes are spread evenly
    (every company flips constantly - possible classifier noise near
    sign boundaries) or concentrated in a subset of companies (more
    consistent with real, legitimate volatility)."""
    by_company: dict[str, list[tuple[str, str]]] = {}
    for r in rows:
        by_company.setdefault(r["company_id"], []).append(
            (r["year"], r["pattern_label"])
        )

    change_counts = {}
    for cid, year_labels in by_company.items():
        year_labels.sort(key=lambda x: x[0])
        n_changes = sum(
            1 for (_, l1), (_, l2) in zip(year_labels, year_labels[1:]) if l1 != l2
        )
        n_transitions = len(year_labels) - 1
        change_counts[cid] = (n_changes, n_transitions)

    print("\n--- Change-frequency diagnostic ---")
    rates = sorted(change_counts.items(), key=lambda x: -(x[1][0] / max(x[1][1], 1)))
    print("Top 10 companies by change RATE (changes / possible transitions):")
    for cid, (n_changes, n_transitions) in rates[:10]:
        rate = n_changes / n_transitions if n_transitions else 0
        print(
            f"  {cid:15s} {n_changes}/{n_transitions} transitions changed ({rate:.0%})"
        )

    print("\nBottom 10 (most STABLE companies):")
    for cid, (n_changes, n_transitions) in rates[-10:]:
        rate = n_changes / n_transitions if n_transitions else 0
        print(
            f"  {cid:15s} {n_changes}/{n_transitions} transitions changed ({rate:.0%})"
        )

    all_rates = [n / max(t, 1) for n, t in change_counts.values()]
    avg_rate = sum(all_rates) / len(all_rates)
    print(f"\nAverage per-company change rate: {avg_rate:.0%}")

    # Distribution of rates — tells us if this is uniform (noise) or
    # bimodal/skewed (a mix of stable and genuinely volatile companies).
    buckets = {"0-20%": 0, "20-40%": 0, "40-60%": 0, "60-80%": 0, "80-100%": 0}
    for rate in all_rates:
        if rate < 0.2:
            buckets["0-20%"] += 1
        elif rate < 0.4:
            buckets["20-40%"] += 1
        elif rate < 0.6:
            buckets["40-60%"] += 1
        elif rate < 0.8:
            buckets["60-80%"] += 1
        else:
            buckets["80-100%"] += 1
    print("\nDistribution of per-company change rates:")
    for bucket, count in buckets.items():
        print(f"  {bucket:10s} {count} companies")

    # Spot check: does GRASIM/M&M/TVSMOTOR/BHEL (Day 31 distress finding)
    # show a recent transition consistent with that finding?
    print("\n--- Spot-check Day 31's flagged companies ---")
    for cid in ["GRASIM", "M&M", "TVSMOTOR", "BHEL"]:
        if cid in by_company:
            history = sorted(by_company[cid], key=lambda x: x[0])
            print(f"  {cid}: {[label for _, label in history[-4:]]}  (last 4 years)")
        else:
            print(f"  {cid}: not found in capital_allocation.csv")


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    """Write csv."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    logger.info("Wrote %d rows -> %s", len(rows), path)


def confirm_cashflow_xlsx_has_label_column() -> bool:
    """Day 31 already wrote capital_allocation_label into
    cashflow_intelligence.xlsx by reusing this same CSV — confirm that
    column is actually populated, not silently blank."""
    from openpyxl import load_workbook

    if not CASHFLOW_XLSX.exists():
        logger.warning(
            "%s not found — cannot confirm capital_allocation_label column",
            CASHFLOW_XLSX,
        )
        return False

    wb = load_workbook(CASHFLOW_XLSX)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "capital_allocation_label" not in header:
        logger.error(
            "capital_allocation_label column NOT found in cashflow_intelligence.xlsx header"
        )
        return False

    col_idx = header.index("capital_allocation_label") + 1
    non_null = sum(
        1
        for row in ws.iter_rows(min_row=2, values_only=False)
        if row[col_idx - 1].value is not None
    )
    total = ws.max_row - 1
    logger.info(
        "capital_allocation_label column: %d/%d rows populated", non_null, total
    )
    return non_null > 0


def main() -> None:
    """Main."""
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found at {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = load_capital_allocation()
        coverage = verify_coverage(conn, rows)

        distribution = latest_year_distribution(rows)
        write_csv(
            OUTPUT_DIR / "capital_allocation_distribution.csv",
            distribution,
            ["pattern_label", "company_count"],
        )

        changes = detect_pattern_changes(rows)
        write_csv(
            OUTPUT_DIR / "pattern_changes.csv",
            changes,
            ["company_id", "from_year", "to_year", "from_pattern", "to_pattern"],
        )

        label_column_ok = confirm_cashflow_xlsx_has_label_column()

        print("\n=== Day 32 Summary ===")
        print("--- Coverage verification ---")
        print(
            f"Companies in DB:                    {coverage['total_companies_in_db']}"
        )
        print(
            f"Companies in capital_allocation.csv: {coverage['companies_in_capital_allocation']}"
        )
        print(f"Missing companies:                   {coverage['missing_companies']}")
        print(f"Cashflow (company,year) rows:         {coverage['cashflow_rows']}")
        print(
            f"Capital allocation (company,year) rows: {coverage['capital_allocation_rows']}"
        )
        print(
            f"Missing (company,year) pairs:          {len(coverage['missing_company_year_pairs'])}"
        )
        if coverage["missing_company_year_pairs"]:
            print(f"  First 10: {coverage['missing_company_year_pairs'][:10]}")

        print("\n--- Latest-year pattern distribution ---")
        for d in distribution:
            print(f"  {d['pattern_label']:25s} {d['company_count']}")

        print("\n--- Pattern changes ---")
        print(f"Total year-over-year pattern changes detected: {len(changes)}")
        print(
            f"Companies with >=1 change: {len(set(c['company_id'] for c in changes))}"
        )

        diagnose_change_frequency(rows)

        print("\n--- cashflow_intelligence.xlsx capital_allocation_label column ---")
        print(f"Populated correctly: {label_column_ok}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
