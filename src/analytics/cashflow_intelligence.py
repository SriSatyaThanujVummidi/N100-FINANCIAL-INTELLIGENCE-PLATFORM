"""
src/analytics/cashflow_intelligence.py

Day 31 — Sprint 5, Module 7 (Cash Flow Intelligence)

Outputs:
    output/cashflow_intelligence.xlsx
        company_id, sector, cfo_quality_score, cfo_quality_label,
        capex_intensity_pct, capex_label, fcf_cagr_5yr, fcf_conversion_pct,
        distress_flag, deleveraging_flag, capital_allocation_label
    output/distress_alerts.csv
        company_id, cfo, cff, net_profit

=== Documented judgment calls ===

1. Built as a NEW module, not a rewrite of Sprint 2 Day 11's signed-off
   src/analytics/cashflow_kpis.py. That file's exact function signatures
   aren't available in this session, so importing/reusing it blind risks
   silently breaking tested code. All metrics here are computed fresh
   from raw tables. FLAG FOR RECONCILIATION: once this runs, compare its
   CFO-quality-score / CapEx-intensity output against cashflow_kpis.py's
   equivalents for a sample of companies — if they diverge, that's a
   real finding to resolve, not something to average away.

2. capital_allocation_label is NOT recomputed here — it's read directly
   from Day 11's already-verified output/capital_allocation.csv (latest
   year per company), avoiding a second, possibly-drifting implementation
   of the 8-pattern CFO/CFI/CFF sign classifier.

3. FCF 5yr CAGR is computed locally (fresh implementation, same edge-case
   family as Day 10's cagr.py / Day 29's parser.py: zero base, sign-flip
   in either direction -> None) rather than importing cagr.py, for the
   same "signature not confirmed" reason as point 1.

4. CFO Quality Score / CapEx Intensity / FCF Conversion are computed from
   raw cashflow + profitandloss tables (joined on company_id, year),
   not read from financial_ratios, since this metric's exact persisted
   column names and coverage there aren't confirmed for this session.

5. Distress flag (CFO<0 AND CFF>0) and Deleveraging flag (CFF<0 AND
   borrowings declining YoY) use each company's own latest reported year
   and, for deleveraging, the year immediately prior. SBIN (Sprint 1
   Day 6: zero balancesheet rows) will correctly show
   deleveraging_flag=None, not False — "unknown" is not the same as "no".
"""

import csv
import logging
import sqlite3
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.etl.fiscal_calendar import get_annual_rows

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"
CAPITAL_ALLOC_CSV = OUTPUT_DIR / "capital_allocation.csv"

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


def get_connection() -> sqlite3.Connection:
    """Get connection."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def shift_fiscal_year(year_str: str, years_back: int) -> str:
    """Shift fiscal year."""
    yyyy, mm = year_str.split("-")
    return f"{int(yyyy) - years_back:04d}-{mm}"


# --------------------------------------------------------------------------
# CFO Quality Score
# --------------------------------------------------------------------------


def cfo_quality_label(avg_ratio: Optional[float]) -> Optional[str]:
    """Cfo quality label."""
    if avg_ratio is None:
        return None
    if avg_ratio > 1.0:
        return "High Quality"
    if avg_ratio >= 0.5:
        return "Moderate"
    return "Accrual Risk"


def compute_cfo_quality(
    conn: sqlite3.Connection, company_id: str
) -> tuple[Optional[float], Optional[str]]:
    """Compute cfo quality."""
    rows = conn.execute(
        "SELECT cf.year, cf.operating_activity, pl.net_profit "
        "FROM cashflow cf JOIN profitandloss pl "
        "ON cf.company_id = pl.company_id AND cf.year = pl.year "
        "WHERE cf.company_id = ? ORDER BY cf.year DESC LIMIT 5",
        (company_id,),
    ).fetchall()

    ratios = []
    for r in rows:
        cfo, pat = r["operating_activity"], r["net_profit"]
        if cfo is None or pat is None or pat == 0:
            continue  # skip, don't zero — Day 11's documented convention
        ratios.append(cfo / pat)

    if not ratios:
        return None, None
    avg_ratio = sum(ratios) / len(ratios)
    return round(avg_ratio, 3), cfo_quality_label(avg_ratio)


# --------------------------------------------------------------------------
# CapEx Intensity
# --------------------------------------------------------------------------


def capex_label(pct: Optional[float]) -> Optional[str]:
    """Capex label."""
    if pct is None:
        return None
    if pct < 3:
        return "Asset Light"
    if pct <= 8:
        return "Moderate"
    return "Capital Intensive"


def compute_capex_intensity(
    conn: sqlite3.Connection, company_id: str
) -> tuple[Optional[float], Optional[str]]:
    """Compute capex intensity."""
    row = conn.execute(
        "SELECT cf.investing_activity, pl.sales "
        "FROM cashflow cf JOIN profitandloss pl "
        "ON cf.company_id = pl.company_id AND cf.year = pl.year "
        "WHERE cf.company_id = ? ORDER BY cf.year DESC LIMIT 1",
        (company_id,),
    ).fetchone()

    if (
        row is None
        or row["investing_activity"] is None
        or row["sales"] is None
        or row["sales"] <= 0
    ):
        return None, None
    pct = abs(row["investing_activity"]) / row["sales"] * 100
    return round(pct, 2), capex_label(pct)


# --------------------------------------------------------------------------
# FCF 5-year CAGR
# --------------------------------------------------------------------------


def compute_fcf_5yr_cagr(conn: sqlite3.Connection, company_id: str) -> Optional[float]:
    """Compute fcf 5yr cagr."""
    latest_row = conn.execute(
        "SELECT MAX(year) AS latest FROM cashflow WHERE company_id = ?", (company_id,)
    ).fetchone()
    if latest_row is None or latest_row["latest"] is None:
        return None
    end_year = latest_row["latest"]
    start_year = shift_fiscal_year(end_year, 5)

    def get_fcf(year: str) -> Optional[float]:
        """Get fcf."""
        r = conn.execute(
            "SELECT operating_activity, investing_activity FROM cashflow "
            "WHERE company_id = ? AND year = ?",
            (company_id, year),
        ).fetchone()
        if (
            r is None
            or r["operating_activity"] is None
            or r["investing_activity"] is None
        ):
            return None
        return r["operating_activity"] + r["investing_activity"]

    end_val = get_fcf(end_year)
    base_val = get_fcf(start_year)

    if base_val is None or end_val is None or base_val == 0:
        return None
    if base_val > 0 and end_val < 0:
        return None  # DECLINE_TO_LOSS
    if base_val < 0 and end_val > 0:
        return None  # TURNAROUND
    if base_val < 0 and end_val < 0:
        return None  # BOTH_NEGATIVE

    return round(((end_val / base_val) ** (1 / 5) - 1) * 100, 2)


# --------------------------------------------------------------------------
# FCF Conversion Rate
# --------------------------------------------------------------------------


def compute_fcf_conversion(
    conn: sqlite3.Connection, company_id: str
) -> Optional[float]:
    """Compute fcf conversion."""
    row = conn.execute(
        "SELECT cf.operating_activity, cf.investing_activity, pl.operating_profit "
        "FROM cashflow cf JOIN profitandloss pl "
        "ON cf.company_id = pl.company_id AND cf.year = pl.year "
        "WHERE cf.company_id = ? ORDER BY cf.year DESC LIMIT 1",
        (company_id,),
    ).fetchone()

    if row is None:
        return None
    cfo, cfi, op_profit = (
        row["operating_activity"],
        row["investing_activity"],
        row["operating_profit"],
    )
    if cfo is None or cfi is None or op_profit is None or op_profit <= 0:
        return None
    fcf = cfo + cfi
    return round(fcf / op_profit * 100, 2)


# --------------------------------------------------------------------------
# Distress & Deleveraging flags
# --------------------------------------------------------------------------


def compute_distress_and_deleveraging(
    conn: sqlite3.Connection, company_id: str, sector: Optional[str] = None
) -> tuple[
    Optional[bool], Optional[bool], Optional[float], Optional[float], Optional[float]
]:
    """Returns (distress_flag, deleveraging_flag, latest_cfo, latest_cff, latest_net_profit).

    Day 31 real-data fix: Financials excluded from distress_flag. Negative
    CFO + positive CFF is the NORMAL cash-flow shape for a lender (loan
    disbursements = CFO outflow, deposit/borrowing raises = CFF inflow),
    not distress — confirmed via day31_diagnose_distress.py: all 9
    Financials flagged under the naive rule posted strongly positive net
    profit (AXISBANK +24,861 Cr, PFC +26,461 Cr, etc.), the opposite of
    what genuine distress would show. Same carve-out family as CON01/
    CON11/D-E/ROCE. deleveraging_flag is unaffected — declining debt is
    a meaningful signal for lenders too, not excluded.
    """
    latest_cf = conn.execute(
        "SELECT year, operating_activity, financing_activity FROM cashflow "
        "WHERE company_id = ? ORDER BY year DESC LIMIT 1",
        (company_id,),
    ).fetchone()

    if latest_cf is None:
        return None, None, None, None, None

    cfo, cff = latest_cf["operating_activity"], latest_cf["financing_activity"]

    distress_flag = None
    if sector != "Financials" and cfo is not None and cff is not None:
        distress_flag = bool(cfo < 0 and cff > 0)

    # Day 33 fix: restrict to this company's own dominant fiscal month —
    # a naive "ORDER BY year DESC LIMIT 2" was silently pairing an
    # off-cycle interim row (e.g. 2024-09) against the last real annual
    # close for ~79 companies, corrupting the year-over-year comparison.
    borrowings_rows = get_annual_rows(
        conn, company_id, "balancesheet", "year, borrowings", limit=2
    )

    deleveraging_flag = None
    if cff is not None and len(borrowings_rows) == 2:
        latest_debt, prior_debt = (
            borrowings_rows[0]["borrowings"],
            borrowings_rows[1]["borrowings"],
        )
        if latest_debt is not None and prior_debt is not None:
            deleveraging_flag = bool(cff < 0 and latest_debt < prior_debt)

    latest_pl = conn.execute(
        "SELECT net_profit FROM profitandloss WHERE company_id = ? "
        "ORDER BY year DESC LIMIT 1",
        (company_id,),
    ).fetchone()
    net_profit = latest_pl["net_profit"] if latest_pl else None

    return distress_flag, deleveraging_flag, cfo, cff, net_profit


# --------------------------------------------------------------------------
# Capital allocation label — reused from Day 11's verified CSV
# --------------------------------------------------------------------------


def load_capital_allocation_labels() -> dict[str, str]:
    """Latest-year pattern_label per company, from Day 11's
    output/capital_allocation.csv (not recomputed here)."""
    if not CAPITAL_ALLOC_CSV.exists():
        logger.warning(
            "capital_allocation.csv not found — capital_allocation_label will be None for all companies"
        )
        return {}

    latest_by_company: dict[str, tuple[str, str]] = {}  # company_id -> (year, label)
    with open(CAPITAL_ALLOC_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid, year, label = row["company_id"], row["year"], row["pattern_label"]
            if cid not in latest_by_company or year > latest_by_company[cid][0]:
                latest_by_company[cid] = (year, label)

    return {cid: label for cid, (_, label) in latest_by_company.items()}


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------


def build_row(
    conn: sqlite3.Connection,
    company_id: str,
    sector: Optional[str],
    alloc_labels: dict[str, str],
) -> dict:
    """Build row."""
    cfo_score, cfo_label = compute_cfo_quality(conn, company_id)
    capex_pct, capex_lbl = compute_capex_intensity(conn, company_id)
    fcf_cagr = compute_fcf_5yr_cagr(conn, company_id)
    fcf_conv = compute_fcf_conversion(conn, company_id)
    distress, deleveraging, cfo, cff, net_profit = compute_distress_and_deleveraging(
        conn, company_id, sector
    )
    ...
    return {
        "company_id": company_id,
        "sector": sector,
        "cfo_quality_score": cfo_score,
        "cfo_quality_label": cfo_label,
        "capex_intensity_pct": capex_pct,
        "capex_label": capex_lbl,
        "fcf_cagr_5yr": fcf_cagr,
        "fcf_conversion_pct": fcf_conv,
        "distress_flag": distress,
        "deleveraging_flag": deleveraging,
        "capital_allocation_label": alloc_labels.get(company_id),
        "_cfo": cfo,
        "_cff": cff,
        "_net_profit": net_profit,
    }


def write_xlsx(rows: list[dict], path: Path) -> None:
    """Write xlsx."""
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = [
        "company_id",
        "sector",
        "cfo_quality_score",
        "cfo_quality_label",
        "capex_intensity_pct",
        "capex_label",
        "fcf_cagr_5yr",
        "fcf_conversion_pct",
        "distress_flag",
        "deleveraging_flag",
        "capital_allocation_label",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Intelligence"

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    distress_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    deleveraging_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    distress_col = columns.index("distress_flag") + 1
    deleveraging_col = columns.index("deleveraging_flag") + 1

    for row in rows:
        ws.append([row[c] for c in columns])
        excel_row = ws.max_row
        if row["distress_flag"] is True:
            ws.cell(row=excel_row, column=distress_col).fill = distress_fill
        if row["deleveraging_flag"] is True:
            ws.cell(row=excel_row, column=deleveraging_col).fill = deleveraging_fill

    for col_cells in ws.columns:
        max_len = max(
            len(str(c.value)) if c.value is not None else 0 for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(30, max_len + 2)

    wb.save(path)
    logger.info("Wrote %d rows -> %s", len(rows), path)


def write_distress_alerts(rows: list[dict], path: Path) -> None:
    """Write distress alerts."""
    flagged = [r for r in rows if r["distress_flag"] is True]
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["company_id", "cfo", "cff", "net_profit"]
        )
        writer.writeheader()
        for r in flagged:
            writer.writerow(
                {
                    "company_id": r["company_id"],
                    "cfo": r["_cfo"],
                    "cff": r["_cff"],
                    "net_profit": r["_net_profit"],
                }
            )
    logger.info("Wrote %d distress-flagged rows -> %s", len(flagged), path)


def main() -> None:
    """Main."""
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found at {DB_PATH}")

    conn = get_connection()
    try:
        company_ids = sorted(r["id"] for r in conn.execute("SELECT id FROM companies"))
        sectors = {
            r["company_id"]: r["broad_sector"]
            for r in conn.execute("SELECT company_id, broad_sector FROM sectors")
        }
        alloc_labels = load_capital_allocation_labels()

        rows = [
            build_row(conn, cid, sectors.get(cid), alloc_labels) for cid in company_ids
        ]

        write_xlsx(rows, OUTPUT_DIR / "cashflow_intelligence.xlsx")
        write_distress_alerts(rows, OUTPUT_DIR / "distress_alerts.csv")

        none_counts = {
            "cfo_quality_score": sum(1 for r in rows if r["cfo_quality_score"] is None),
            "capex_intensity_pct": sum(
                1 for r in rows if r["capex_intensity_pct"] is None
            ),
            "fcf_cagr_5yr": sum(1 for r in rows if r["fcf_cagr_5yr"] is None),
            "fcf_conversion_pct": sum(
                1 for r in rows if r["fcf_conversion_pct"] is None
            ),
            "distress_flag": sum(1 for r in rows if r["distress_flag"] is None),
            "deleveraging_flag": sum(1 for r in rows if r["deleveraging_flag"] is None),
            "capital_allocation_label": sum(
                1 for r in rows if r["capital_allocation_label"] is None
            ),
        }

        distress_count = sum(1 for r in rows if r["distress_flag"] is True)
        deleveraging_count = sum(1 for r in rows if r["deleveraging_flag"] is True)

        cfo_label_dist: dict[str, int] = {}
        capex_label_dist: dict[str, int] = {}
        for r in rows:
            if r["cfo_quality_label"]:
                cfo_label_dist[r["cfo_quality_label"]] = (
                    cfo_label_dist.get(r["cfo_quality_label"], 0) + 1
                )
            if r["capex_label"]:
                capex_label_dist[r["capex_label"]] = (
                    capex_label_dist.get(r["capex_label"], 0) + 1
                )

        print("\n=== Day 31 Summary ===")
        print(f"Companies processed: {len(rows)}")
        print(
            "\nNull counts per column (companies where this metric couldn't be computed):"
        )
        for col, count in none_counts.items():
            print(f"  {col}: {count}")
        print(f"\nCFO quality label distribution: {cfo_label_dist}")
        print(f"CapEx label distribution: {capex_label_dist}")
        print(f"\nDistress flagged: {distress_count}")
        print(f"Deleveraging flagged: {deleveraging_count}")

        if distress_count > 0:
            print("\nDistress companies:")
            for r in rows:
                if r["distress_flag"] is True:
                    print(
                        f"  {r['company_id']}: CFO={r['_cfo']}, CFF={r['_cff']}, NetProfit={r['_net_profit']}"
                    )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
