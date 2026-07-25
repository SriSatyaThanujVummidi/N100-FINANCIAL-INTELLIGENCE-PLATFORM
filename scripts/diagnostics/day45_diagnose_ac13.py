"""Day 45 diagnostic -- redo AC-13 with the FULL Quality Compounder threshold set (my first
pass omitted min_rev_cagr_5yr), and separately check whether ICICIBANK/CANBK's absence from
the API result is explained by a missing D/E-Financials carve-out in src/api/routers/screener.py."""
import sqlite3
import requests

API_BASE = "http://localhost:8000/api/v1"

# Full Quality Compounder preset: ROE>15%, D/E<1 (Financials exempt), FCF>0, Rev CAGR 5yr>10%
resp = requests.get(f"{API_BASE}/screener", params={
    "min_roe": 15, "max_de": 1, "min_fcf": 0, "min_rev_cagr_5yr": 10,
})
api_companies = {c["company_id"] for c in resp.json()["results"]}

import openpyxl
wb = openpyxl.load_workbook("output/screener_output.xlsx", data_only=True)
sheet_name = next((s for s in wb.sheetnames if "quality" in s.lower()), wb.sheetnames[0])
ws = wb[sheet_name]
headers = [c.value for c in ws[1]]
id_col = headers.index("company_id") if "company_id" in headers else 0
xlsx_companies = {row[id_col].value for row in ws.iter_rows(min_row=2) if row[id_col].value}

print(f"API (all 4 filters):   {len(api_companies)} companies")
print(f"Excel (quality_compounder): {len(xlsx_companies)} companies")
print(f"Match: {api_companies == xlsx_companies}")
print(f"API only: {api_companies - xlsx_companies}")
print(f"Excel only: {xlsx_companies - api_companies}")

# Check ICICIBANK/CANBK specifically -- are they Financials with D/E > 1?
conn = sqlite3.connect("data/nifty100.db")
for ticker in ["ICICIBANK", "CANBK"]:
    row = conn.execute(
        """SELECT fr.debt_to_equity, s.broad_sector FROM financial_ratios fr
           JOIN sectors s ON fr.company_id = s.company_id
           WHERE fr.company_id=? ORDER BY fr.year DESC LIMIT 1""", (ticker,)
    ).fetchone()
    print(f"{ticker}: D/E={row[0]}, sector={row[1]}")
conn.close()